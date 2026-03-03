#!/usr/bin/env python3
"""
银行日记账自动生成工具
从招商银行网银HTML交易明细自动生成Excel日记账条目

用法:
  python3 journal_generator.py --html bank.html --excel 日记账.xlsx --month 2
  python3 journal_generator.py --clipboard --excel 日记账.xlsx --month 2
"""

import argparse
import re
import sys
import subprocess
from copy import copy

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("错误: 需要安装 beautifulsoup4")
    print("运行: pip install beautifulsoup4")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl")
    print("运行: pip install openpyxl")
    sys.exit(1)


# ============================================================
# 分类规则配置 — 根据实际情况修改
# ============================================================

# 外包公司名称 → 备注标签 (留空表示不加备注)
OUTSOURCE_COMPANIES = {
    '菏泽上禾数字科技有限公司': '',
    '黄岛区萧氏探索互联网信息科技家（个体工商户）': 'xiao',
    '深圳市汇思艺工业设计有限公司': '',
    # 可继续添加...
}

# 税单号前缀 → 税费类型
# "437" 开头 = 社保(保险); "626" 开头 = 个人所得税
# 遇到新的前缀时在此添加
TAX_PREFIX_MAP = {
    '626': '个人所得税',
    '437': '保险',          # 自动区分: 金额小的=基本医疗保险, 大的=保险
}


# ============================================================
# HTML 解析
# ============================================================

def parse_reg_time(time_str):
    """
    解析登记时间字符串 → (月, 日)
    支持格式: "2026-02-07 10:30:00", "2026/02/07", "20260207", "02-07" 等
    """
    if not time_str:
        return None
    # 格式: YYYY-MM-DD 或 YYYY/MM/DD (可能带时间)
    m = re.match(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', time_str)
    if m:
        return (int(m.group(2)), int(m.group(3)))
    # 格式: YYYYMMDD
    m = re.match(r'(\d{4})(\d{2})(\d{2})', time_str)
    if m:
        return (int(m.group(2)), int(m.group(3)))
    # 格式: MM-DD 或 MM/DD
    m = re.match(r'(\d{1,2})[-/](\d{1,2})', time_str)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    return None


def parse_bank_html(html_str):
    """解析招商银行网银HTML表格，提取交易数据列表"""
    soup = BeautifulSoup(html_str, 'html.parser')
    transactions = []

    field_map = {
        'dTrsDay': 'reg_time',           # 登记时间
        'cTrsMod': 'direction',          # 出账/入账
        'cCcyNbr': 'currency',           # 币种
        'zTrsAmt': 'amount_str',         # 金额(字符串)
        'zCtpEac': 'counterparty_acct',  # 对方账号
        'zCtpNam': 'counterparty_name',  # 对方名称
        'zTrsTxt': 'summary',            # 摘要
        'cBusTyp': 'business_type',      # 业务类型
        'cPrtTag': 'print_status',       # 打印状态
    }

    # 登记时间字段可能使用不同的 data-field 名称
    date_field_candidates = ['dTrsDay', 'dRegDat', 'dTrsDat', 'dRegDay', 'zRegDatTim']

    for row in soup.find_all('tr', class_='ant-table-row'):
        tx = {}
        for cell in row.find_all('td', attrs={'data-field': True}):
            field = cell['data-field']
            # 检查是否为日期字段
            if field in date_field_candidates:
                text = cell.get_text(strip=True)
                tx['reg_time'] = text
                continue
            if field not in field_map:
                continue
            # Ant Design 表格的文本在嵌套 span 中
            text = cell.get_text(strip=True)
            tx[field_map[field]] = text

        if 'amount_str' in tx:
            tx['amount'] = float(tx['amount_str'].replace(',', ''))
            # 解析登记时间 → (月, 日)
            if tx.get('reg_time'):
                tx['parsed_date'] = parse_reg_time(tx['reg_time'])
            transactions.append(tx)

    return transactions


# ============================================================
# 交易分类
# ============================================================

def extract_date(summary):
    """从摘要中提取日期 (YYYYMMDD 格式) → (月, 日) 或 None"""
    match = re.search(r'(20\d{6})', summary or '')
    if match:
        d = match.group(1)
        month, day = int(d[4:6]), int(d[6:8])
        if 1 <= month <= 12 and 1 <= day <= 31:
            return (month, day)
    return None


def extract_tax_number(summary):
    """从摘要中提取税单号"""
    match = re.search(r'税单号[：:](\d+)', summary or '')
    return match.group(1) if match else ''


def classify(tx):
    """
    对单笔交易分类
    返回: {category, description, io_type, date, note, group_tag}
    """
    direction = tx.get('direction', '')
    amount = tx.get('amount', 0)
    summary = tx.get('summary', '-')
    biz_type = tx.get('business_type', '-')
    cpty_name = tx.get('counterparty_name', '-')

    io_type = 'income' if direction == '入账' else 'expense'
    # 优先使用HTML中的登记时间，回退到摘要中的日期
    date = tx.get('parsed_date') or extract_date(summary)

    # ---- 银行服务费 ----
    if biz_type == '企业银行收费':
        return dict(category='服务费', description='银行服务费',
                    io_type='expense', date=date, note='', group_tag=None)

    # ---- 工资 ----
    if biz_type == '自助代发付款' or summary == '工资':
        return dict(category='工资', description='工资发放',
                    io_type='expense', date=date, note='', group_tag=None)

    # ---- 备用金(入账) ----
    if direction == '入账' and '备用金' in summary:
        return dict(category='备用金', description='总部拨付备用金',
                    io_type='income', date=date, note='', group_tag=None)

    # ---- 利息(入账) ----
    if direction == '入账' and '利息' in summary:
        return dict(category='利息', description='利息',
                    io_type='income', date=date, note='', group_tag=None)

    # ---- 公积金 ----
    if '公积金' in (cpty_name or ''):
        return dict(category='公积金', description='公积金',
                    io_type=io_type, date=date, note='', group_tag=None)

    # ---- 缴税(个人所得税/保险/企业所得税/公积金 等) ----
    if '实时缴税' in (summary or ''):
        tax_num = extract_tax_number(summary)
        prefix = tax_num[:3] if len(tax_num) >= 3 else ''
        tax_type = TAX_PREFIX_MAP.get(prefix)

        if tax_type == '个人所得税':
            return dict(category='个人所得税', description='个人所得税',
                        io_type='expense', date=date, note='', group_tag=None)
        elif tax_type == '保险':
            # 先标记为 insurance, 后续再区分基本医疗保险 vs 保险
            return dict(category='保险', description='待区分',
                        io_type='expense', date=date, note='', group_tag='insurance')
        else:
            return dict(category='税费', description=f'缴税(前缀{prefix})',
                        io_type='expense', date=date, note='', group_tag=None)

    # ---- 外包费(移动支付给公司) ----
    if biz_type == '移动支付' and cpty_name not in ['-', '', None]:
        note = OUTSOURCE_COMPANIES.get(cpty_name, '')
        return dict(category='外包费', description='外包服务费',
                    io_type='expense', date=date, note=note, group_tag=None)

    # ---- 其他入账 ----
    if direction == '入账':
        desc = cpty_name if cpty_name not in ['-', ''] else summary
        return dict(category='其他收入', description=desc,
                    io_type='income', date=date, note='', group_tag=None)

    # ---- 未分类 ----
    return dict(category='未分类', description=summary or biz_type,
                io_type='expense', date=date, note='', group_tag=None)


# ============================================================
# 生成日记账条目
# ============================================================

def process_transactions(transactions, month):
    """处理交易列表，生成日记账条目"""

    # 0. 按日期排序(不过滤月份)
    sorted_txs = sorted(transactions, key=lambda tx: (
        tx.get('parsed_date', (month, 99))[0],
        tx.get('parsed_date', (month, 99))[1]
    ))

    # 1. 分类每笔交易
    items = []
    for tx in sorted_txs:
        cls = classify(tx)
        cls['amount'] = tx['amount']
        cls['counterparty'] = tx.get('counterparty_name', '')
        items.append(cls)

    result = list(items)

    # 2. 区分保险类型: 金额小 → 基本医疗保险, 金额大 → 保险
    ins_indices = [i for i, item in enumerate(result) if item.get('group_tag') == 'insurance']
    ins_items = [result[i] for i in ins_indices]
    if len(ins_items) >= 2:
        ins_items.sort(key=lambda x: x['amount'])
        ins_items[0]['description'] = '基本医疗保险'
        for item in ins_items[1:]:
            item['description'] = '保险'
        for j, idx in enumerate(ins_indices):
            result[idx] = ins_items[j]
    elif len(ins_items) == 1:
        ins_items[0]['description'] = '保险'

    # 3. 清除临时标记
    for item in result:
        item.pop('group_tag', None)

    # 4. 生成凭证号(银1, 银2, 银3, ... 顺序编号)
    entries = []
    for num, item in enumerate(result, 1):
        date = item.get('date')
        entry = {
            'A': f'银{num}',                    # 凭证号
            'B': date[0] if date else month,     # 月
            'C': date[1] if date else None,      # 日
            'D': item['category'],               # 费用分类
            'E': item['description'],            # 摘要
            'H': item['amount'] if item['io_type'] == 'income' else None,
            'I': item['amount'] if item['io_type'] == 'expense' else None,
            'note': item.get('note', ''),        # 备注
        }
        entries.append(entry)

        num += 1
        i = j

    return entries


# ============================================================
# 写入 Excel
# ============================================================

def detect_balance_col(ws):
    """检测余额列(J 或 K)"""
    for col in ['J', 'K']:
        val = ws[f'{col}1'].value
        if val and '余额' in str(val):
            return col
    # 默认用 J
    return 'J'


def find_last_ytd_row(ws):
    """找到最后一个"本年累计"行号"""
    last = None
    for row in range(3, ws.max_row + 1):
        if ws[f'E{row}'].value == '本年累计':
            last = row
    return last


def find_ref_rows(ws):
    """找到第一个月的数据行、本月合计行、本年累计行(作为格式参考)"""
    data_row = total_row = ytd_row = None
    for row in range(3, ws.max_row + 1):
        e_val = ws[f'E{row}'].value
        if e_val == '本月合计' and total_row is None:
            total_row = row
        elif e_val == '本年累计' and ytd_row is None:
            ytd_row = row
        elif ws[f'A{row}'].value and str(ws[f'A{row}'].value).startswith('银') and data_row is None:
            data_row = row
    return data_row or 4, total_row or 15, ytd_row or 16


def copy_cell_format(ws, src_row, dst_row, columns):
    """复制一行的单元格格式"""
    for col in columns:
        src = ws[f'{col}{src_row}']
        dst = ws[f'{col}{dst_row}']
        if src.font:
            dst.font = copy(src.font)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.border:
            dst.border = copy(src.border)
        if src.number_format:
            dst.number_format = src.number_format
        if src.fill:
            dst.fill = copy(src.fill)


def write_to_excel(entries, excel_path, sheet_name='银行日记账2025'):
    """将日记账条目写入Excel文件"""
    wb = openpyxl.load_workbook(excel_path)

    if sheet_name not in wb.sheetnames:
        print(f"错误: 找不到工作表 '{sheet_name}'")
        print(f"可用的工作表: {wb.sheetnames}")
        return None

    ws = wb[sheet_name]
    bal_col = detect_balance_col(ws)
    note_col = chr(ord(bal_col) + 1)  # 备注列 = 余额列的下一列

    # 找到上月最后行
    last_ytd = find_last_ytd_row(ws)
    if last_ytd is None:
        print("错误: 找不到上月的'本年累计'行")
        return None

    # 找格式参考行
    ref_data, ref_total, ref_ytd = find_ref_rows(ws)

    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', bal_col]
    start_row = last_ytd + 1
    last_data_row = start_row + len(entries) - 1
    total_row = last_data_row + 1
    ytd_row = total_row + 1

    # 写入数据行
    for idx, entry in enumerate(entries):
        row = start_row + idx
        copy_cell_format(ws, ref_data, row, columns)

        ws[f'A{row}'] = entry['A']
        ws[f'B{row}'] = entry['B']
        if entry['C'] is not None:
            ws[f'C{row}'] = entry['C']
        ws[f'D{row}'] = entry['D']
        ws[f'E{row}'] = entry['E']
        if entry['H'] is not None:
            ws[f'H{row}'] = entry['H']
        if entry['I'] is not None:
            ws[f'I{row}'] = entry['I']
        if entry.get('note'):
            ws[f'{note_col}{row}'] = entry['note']

        # 余额公式
        prev = last_ytd if row == start_row else row - 1
        ws[f'{bal_col}{row}'] = f'={bal_col}{prev}+H{row}-I{row}'

    # 本月合计
    copy_cell_format(ws, ref_total, total_row, columns)
    ws[f'E{total_row}'] = '本月合计'
    ws[f'H{total_row}'] = f'=SUM(H{start_row}:H{last_data_row})'
    ws[f'I{total_row}'] = f'=SUM(I{start_row}:I{last_data_row})'
    ws[f'{bal_col}{total_row}'] = f'={bal_col}{last_ytd}+H{total_row}-I{total_row}'

    # 本年累计
    copy_cell_format(ws, ref_ytd, ytd_row, columns)
    ws[f'E{ytd_row}'] = '本年累计'
    ws[f'H{ytd_row}'] = f'={bal_col}{last_ytd}+H{total_row}'
    ws[f'I{ytd_row}'] = f'=I{total_row}'
    ws[f'{bal_col}{ytd_row}'] = f'=H{ytd_row}-I{ytd_row}'

    wb.save(excel_path)

    return {
        'start_row': start_row,
        'last_data_row': last_data_row,
        'total_row': total_row,
        'ytd_row': ytd_row,
        'balance_col': bal_col,
    }


# ============================================================
# 主程序
# ============================================================

def print_preview(entries, month):
    """打印预览表格"""
    print(f"\n生成 {len(entries)} 条日记账条目 ({month}月):\n")
    header = f"{'凭证号':<8} {'月':>2} {'日':>3}  {'费用分类':<10} {'摘要':<12} {'收入':>12} {'支出':>12}  {'备注'}"
    print(header)
    print('-' * len(header.encode('gbk', errors='replace')))

    for e in entries:
        h_str = f"{e['H']:>12,.2f}" if e['H'] else ' ' * 12
        i_str = f"{e['I']:>12,.2f}" if e['I'] else ' ' * 12
        day_str = f"{e['C']}" if e['C'] else '??'
        note = e.get('note', '')
        print(f"{e['A']:<8} {e['B']:>2} {day_str:>3}  {e['D']:<10} {e['E']:<12} {h_str} {i_str}  {note}")

    total_in = sum(e['H'] for e in entries if e['H'])
    total_out = sum(e['I'] for e in entries if e['I'])
    print('-' * 80)
    print(f"{'合计':<16}  {' ':<10} {' ':<12} {total_in:>12,.2f} {total_out:>12,.2f}")


def main():
    parser = argparse.ArgumentParser(
        description='银行日记账自动生成工具 - 从招商银行网银HTML自动写入Excel日记账',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  %(prog)s --html bank_feb.html --excel 日记账202601.xlsx --month 2
  %(prog)s --clipboard --excel 日记账202601.xlsx --month 2

分类规则:
  企业银行收费     → 服务费/银行服务费 (同月合并为一条)
  自助代发付款     → 工资/工资发放
  入账+备用金      → 备用金/总部拨付备用金
  移动支付+公司名  → 外包费/外包服务费
  实时缴税+626前缀 → 个人所得税
  实时缴税+437前缀 → 保险/基本医疗保险 (按金额大小区分)

如需添加新规则，请修改脚本顶部的配置区。
""")

    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument('--html', metavar='FILE', help='HTML文件路径')
    source.add_argument('--clipboard', action='store_true',
                        help='从剪贴板读取HTML (macOS)')

    parser.add_argument('--excel', required=True, metavar='FILE',
                        help='Excel日记账文件路径')
    parser.add_argument('--month', type=int, required=True, choices=range(1, 13),
                        metavar='N', help='月份 (1-12)')
    parser.add_argument('--sheet', default='银行日记账2025', metavar='NAME',
                        help='工作表名称 (默认: 银行日记账2025)')
    parser.add_argument('--dry-run', action='store_true',
                        help='仅预览，不写入Excel')

    args = parser.parse_args()

    # 读取 HTML
    if args.clipboard:
        result = subprocess.run(['pbpaste'], capture_output=True, text=True)
        html_str = result.stdout
        if not html_str.strip():
            print("错误: 剪贴板为空")
            sys.exit(1)
    else:
        with open(args.html, 'r', encoding='utf-8') as f:
            html_str = f.read()

    # 解析交易
    transactions = parse_bank_html(html_str)
    if not transactions:
        print("错误: 未从HTML中解析到任何交易记录")
        print("请确认粘贴的是招商银行网银的交易明细表格HTML")
        sys.exit(1)

    print(f"解析到 {len(transactions)} 笔银行交易")

    # 生成日记账条目
    entries = process_transactions(transactions, args.month)

    # 预览
    print_preview(entries, args.month)

    # 检查缺少日期的条目
    no_date = [e for e in entries if e['C'] is None]
    if no_date:
        print(f"\n注意: 以下 {len(no_date)} 条缺少日期(日列显示??), 需在Excel中手动补填:")
        for e in no_date:
            print(f"  {e['A']:>8}  {e['D']}/{e['E']}")

    # 检查未分类的条目
    unknown = [e for e in entries if e['D'] in ('未分类', '税费')]
    if unknown:
        print(f"\n注意: 以下条目分类可能不准确, 请核实:")
        for e in unknown:
            print(f"  {e['A']:>8}  {e['D']}/{e['E']}")

    if args.dry_run:
        print("\n(--dry-run 模式, 未写入Excel)")
        return

    # 写入 Excel
    info = write_to_excel(entries, args.excel, args.sheet)
    if info:
        print(f"\n已写入: {args.excel}")
        print(f"  数据行: {info['start_row']}-{info['last_data_row']}")
        print(f"  本月合计: 行 {info['total_row']}")
        print(f"  本年累计: 行 {info['ytd_row']}")
        print(f"  余额列: {info['balance_col']}")
    else:
        print("\n写入失败，请检查错误信息")
        sys.exit(1)


if __name__ == '__main__':
    main()
